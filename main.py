# By: Abdulrahman Alamoudi
# Created Date: Nov 21, 2016
# status : does th job

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
from time import gmtime, strftime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import json

def main():

    waitTime = 1800 #1800
    while True:
        try:
            with open("data.json", "r") as file:
                manip = json.load(file)
                manip = FixList(manip)

            # manip is [4, ["time", link, link, link...]]

            wb = load_workbook("data.xlsx")
            if time.strftime("%d-%m-%y") not in wb.sheetnames:
                ws = wb.create_sheet(time.strftime("%d-%m-%y"),0)
                manip[0] = 4
                #ws.title = time.strftime("%x")
            else:
                ws = wb.get_sheet_by_name(time.strftime("%d-%m-%y"))

            rowNum = str(manip[0])  # 5 is the default
                                    #string more is flixible speciely in excel

            driver = webdriver.Chrome()
            for item in range(0, len(manip[1])) :
                link = manip[1][item][0]
                if link != "time":
                    driver.get(link)
                    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//span[@id='priceblock_ourprice'] | //span[@id='priceblock_dealprice'] | //span[@id='priceblock_saleprice']")))

                #if manip[0] == 4:
                excelColumnLetter = manip[1][item][1]
                if str(ws[excelColumnLetter +"1"].value) != "TIME" and excelColumnLetter == "A":
                # initialize time column
                    ws = ExcelInsertTitleAndLink(manip[1][item], driver, ws)
                elif link != str(ws[manip[1][item][1]+"2"].value) and excelColumnLetter != "A":
                # initialize link column
                    ws = CleanColumn(ws, excelColumnLetter, rowNum)     #in case of a previous item exist in column : clean
                    ws = ExcelInsertTitleAndLink(manip[1][item], driver, ws)
                    ws = ExcelInsertPrice(manip[1][item], rowNum, driver, ws)
                elif link == "time":
                    ws = ExcelInsertTime(manip[1][item], rowNum, ws)
                else:
                    ws = ExcelInsertPrice(manip[1][item], rowNum, driver, ws)

            driver.quit()

            wb.save("data.xlsx")

            #compare the most right bottom value with ""
            if str(ws[manip[1][-1][1]+rowNum].value) != "":   #this if state. to don't leave empty rows in between
                manip[0] = manip[0] + 1

            with open("data.json", "w") as file:
                json.dump(manip, file)

            countdown(waitTime)

        except:
            driver.quit()


def GetPrice(driver, html):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    for search in soup.find_all("span", {"id":"priceblock_ourprice"}):
        return search.text

    for search in soup.find_all("span", {"id": "priceblock_dealprice"}):
        return search.text

    for search in soup.find_all("span", {"id": "priceblock_saleprice"}):
        return search.text


def GetName(driver, html):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    for search in soup.find_all("h1", {"id":"title"}):
        return CleanName(search.text)

def CleanName(name):
    ###### removing spaces and new lines from the beginning
    count = 0
    while (name[count] == " " or name[count] == "\n"):
        count += 1
    name = name[count:]

    ###### removing spaces and new lines from the end
    count = len(name) - 1
    while (name[count] == " " or name[count] == "\n"):
        count -= 1
    name = name[:count]

    return name


def GetAZList(numNeeded):   #excel A-Z
    AZLetters = []

    for letter in range(65, 91): #A - Z (65 - 90)
        AZLetters.append(chr(letter))

    excelLetterList = AZLetters

    while True:
        for x in excelLetterList:
            for letter in range(0, 26):
                excelLetterList.append(x + AZLetters[letter])
                if numNeeded <= len(excelLetterList):
                    return excelLetterList

def FixList(manip): #fixing "data.json" by making it 2d list [[link], [excel column], last excel row ]
    listLength = len(manip)
    excelColumnAZ = GetAZList(listLength)

    for x in range(len(manip[1])):
        if len(manip[1][x]) != 2:
            manip[1][x] = [manip[1][x], excelColumnAZ[x]]
    return manip

def ExcelInsertTitleAndLink(item, driver, ws):


    if item[0] == "time":   #time
        ws[item[1] + "1"] = "TIME"
        ws[item[1] + "4"] = time.strftime('%X')
    else:   #item column
        link = item[0]
        html = driver.page_source
        name = GetName(driver, html)
        price = GetPrice(driver, html)

        ws[item[1]+"1"] = name
        ws[item[1]+"1"].alignment = Alignment(wrapText=True)    #wrap Text

        ws[item[1]+"2"] = link

    return ws

def ExcelInsertPrice(item, rowNum, driver, ws):

    html = driver.page_source

    price = GetPrice(driver, html)
    ws[item[1] + rowNum] = price

    return ws

def ExcelInsertTime(item, rowNum, ws):
    ws[item[1] + rowNum] = time.strftime('%X')

    return ws

def countdown(t):
    while t:
        time.sleep(1)
        t -= 1

def CleanColumn(ws, columnCharacter, rowNum):
    for cellNum in range(1, int(rowNum)+1):
        ws[columnCharacter + str(cellNum)] = ""
    return ws

main()
